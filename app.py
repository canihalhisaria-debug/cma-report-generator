import io
from dataclasses import dataclass

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

st.set_page_config(page_title="CMA Dashboard", layout="wide")

# --- Premium CSS ---
st.markdown(
    """
    <style>
        :root {
            --bg: #f3f6fb;
            --card: #ffffff;
            --text: #1b2a41;
            --muted: #64748b;
            --primary: #0f4c81;
            --accent: #2e7ad7;
            --success: #0f766e;
            --border: #e2e8f0;
        }

        .stApp {
            background: radial-gradient(circle at top left, #eef4ff 0%, var(--bg) 45%, #f8fafc 100%);
            color: var(--text);
            font-family: "Inter", "Segoe UI", sans-serif;
        }

        .top-shell {
            background: linear-gradient(135deg, #0f172a 0%, #102a43 40%, #123b74 100%);
            border-radius: 18px;
            padding: 24px 28px;
            box-shadow: 0 14px 28px rgba(15, 23, 42, 0.18);
            margin-bottom: 20px;
        }

        .business-name {
            font-size: 2.2rem;
            font-weight: 800;
            line-height: 1.1;
            background: linear-gradient(90deg, #93c5fd, #e2e8f0, #a7f3d0);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: 0.4px;
            margin-bottom: 6px;
        }

        .business-address {
            color: #dbeafe;
            font-size: 0.98rem;
            letter-spacing: 0.2px;
        }

        .main-heading {
            font-size: 1.28rem;
            font-weight: 700;
            color: #0f172a;
            margin: 6px 0 18px 0;
            padding-left: 6px;
            border-left: 4px solid #2563eb;
        }

        .section-title {
            font-size: 1.08rem;
            font-weight: 700;
            color: #1e3a8a;
            margin: 12px 0 8px 2px;
            letter-spacing: 0.2px;
        }

        .card {
            background: var(--card);
            border: 1px solid var(--border);
            border-radius: 16px;
            padding: 14px 16px 8px;
            margin-bottom: 14px;
            box-shadow: 0 8px 22px rgba(15, 23, 42, 0.08);
        }

        .card-title {
            font-size: 1rem;
            font-weight: 700;
            color: #0f4c81;
            margin-bottom: 10px;
        }

        .summary-box {
            background: linear-gradient(160deg, #ffffff 0%, #f8fbff 100%);
            border: 1px solid #dbe7f7;
            border-radius: 18px;
            box-shadow: 0 10px 24px rgba(15, 23, 42, 0.08);
            padding: 16px;
            position: sticky;
            top: 1rem;
        }

        .summary-title {
            font-size: 1.06rem;
            font-weight: 800;
            color: #0f4c81;
            margin-bottom: 10px;
        }

        .metric-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px dashed #e2e8f0;
            padding: 7px 0;
            font-size: 0.95rem;
        }

        .metric-row:last-child {
            border-bottom: none;
        }

        .metric-label {
            color: #334155;
            font-weight: 500;
        }

        .metric-val {
            color: #0f172a;
            font-weight: 700;
        }

        div[data-testid="stNumberInput"] input,
        div[data-testid="stTextInput"] input,
        div[data-testid="stSelectbox"] div[data-baseweb="select"] {
            border-radius: 10px !important;
        }

        .button-shell {
            background: #ffffff;
            border-radius: 16px;
            border: 1px solid #e2e8f0;
            box-shadow: 0 8px 20px rgba(15, 23, 42, 0.06);
            padding: 16px;
            margin-top: 10px;
        }

        .stButton > button, .stDownloadButton > button {
            width: 100%;
            border-radius: 10px;
            border: 0;
            font-weight: 600;
            padding: 0.55rem 0.8rem;
        }

        .stButton > button[kind="primary"] {
            background: linear-gradient(90deg, #2563eb, #1d4ed8);
            color: white;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


@dataclass
class Summary:
    total_income: float
    material_cost: float
    ebit: float
    total_interest: float
    ebt: float
    pat: float
    cash_accrual: float
    working_capital: float


def num(label: str, key: str) -> float:
    return st.number_input(label, min_value=0.0, value=0.0, step=1000.0, key=key, format="%.2f")


INDIAN_NUMBER_FMT = "#,##,##0.00"
INDIAN_PERCENT_FMT = "0.00"


def build_styled_cma_workbook(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    thin_side = Side(style="thin", color="D1D5DB")
    medium_side = Side(style="medium", color="4B5563")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="0B2545"),
            "fill": PatternFill(fill_type="solid", fgColor="E6EFFA"),
            "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(fill_type="solid", fgColor="1F4E78"),
            "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section": {
            "font": Font(name="Calibri", size=11, bold=True, color="1E3A8A"),
            "fill": PatternFill(fill_type="solid", fgColor="DEEAF6"),
            "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "data": {
            "font": Font(name="Calibri", size=10, color="111827"),
            "fill": PatternFill(fill_type="solid", fgColor="FFFFFF"),
            "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "formula": {
            "font": Font(name="Calibri", size=10, italic=True, color="374151"),
            "fill": PatternFill(fill_type="solid", fgColor="F3F4F6"),
            "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "subtotal": {
            "font": Font(name="Calibri", size=10, bold=True, color="1F2937"),
            "fill": PatternFill(fill_type="solid", fgColor="EAF2FB"),
            "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "total": {
            "font": Font(name="Calibri", size=11, bold=True, color="0F172A"),
            "fill": PatternFill(fill_type="solid", fgColor="D9E5F7"),
            "align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
    }

    def render_sheet(sheet_name: str, report_title: str, rows: list[dict], amount_header: str = "Amount"):
        ws = wb.create_sheet(title=sheet_name)
        ws.merge_cells("A1:B1")
        ws["A1"] = report_title

        ws["A2"] = "Particulars"
        ws["B2"] = amount_header

        ws.column_dimensions["A"].width = 62
        ws.column_dimensions["B"].width = 24
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

        for cell in (ws["A1"], ws["B1"]):
            cell.font = styles["title"]["font"]
            cell.fill = styles["title"]["fill"]
            cell.alignment = styles["title"]["align"]
            cell.border = thin_border

        for cell in (ws["A2"], ws["B2"]):
            cell.font = styles["header"]["font"]
            cell.fill = styles["header"]["fill"]
            cell.alignment = styles["header"]["align"]
            cell.border = thin_border

        row_idx = 3
        for row in rows:
            row_type = row.get("type", "data")
            style = styles.get(row_type, styles["data"])

            label_cell = ws.cell(row=row_idx, column=1, value=row.get("label", ""))
            value_cell = ws.cell(row=row_idx, column=2, value=row.get("formula", row.get("value")))

            label_cell.font = style["font"]
            label_cell.fill = style["fill"]
            label_cell.alignment = style["align"]
            label_cell.border = thin_border

            value_cell.font = style["font"]
            value_cell.fill = style["fill"]
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
            value_cell.number_format = row.get("number_format", INDIAN_NUMBER_FMT)

            if row_type in {"subtotal", "total"}:
                label_cell.border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=thin_side)
                value_cell.border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=thin_side)

            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 24

    pnl_rows = [
        {"type": "section", "label": "A. INCOME"},
        {"type": "data", "label": "Domestic Sale", "value": data["domestic_sale"]},
        {"type": "data", "label": "Export Sale", "value": data["export_sale"]},
        {"type": "data", "label": "Other Income", "value": data["other_income"]},
        {"type": "subtotal", "label": "Total Income (A)", "formula": "=SUM(B4:B6)"},
        {"type": "section", "label": "B. DIRECT / MATERIAL COST"},
        {"type": "data", "label": "Opening Stock", "value": data["opening_stock"]},
        {"type": "data", "label": "Purchases", "value": data["purchases"]},
        {"type": "data", "label": "Carriage Outward", "value": data["carriage_outward"]},
        {"type": "data", "label": "Unloading Expenses", "value": data["unloading_expenses"]},
        {"type": "data", "label": "Direct Expenses", "value": data["direct_expenses"]},
        {"type": "data", "label": "Less: Closing Stock", "value": -data["closing_stock"]},
        {"type": "subtotal", "label": "Material Cost (B)", "formula": "=SUM(B9:B14)"},
        {"type": "section", "label": "C. INDIRECT / OPERATING EXPENSES"},
        {"type": "data", "label": "Salary & Wages", "value": data["salary_wages"]},
        {"type": "data", "label": "Power & Fuel", "value": data["power_fuel"]},
        {"type": "data", "label": "Rent Expenses", "value": data["rent_exp"]},
        {"type": "data", "label": "Printing & Stationery", "value": data["printing_stationery"]},
        {"type": "data", "label": "Depreciation", "value": data["depreciation"]},
        {"type": "data", "label": "Other Expenditure", "value": data["other_expenditure"]},
        {"type": "subtotal", "label": "Operating Expenses (C)", "formula": "=SUM(B17:B22)"},
        {"type": "total", "label": "D. EBIT", "formula": "=B7-B15-B23"},
        {"type": "section", "label": "E. INTEREST"},
        {"type": "data", "label": "Interest on CC", "value": data["interest_cc"]},
        {"type": "data", "label": "Interest on TL", "value": data["interest_tl"]},
        {"type": "subtotal", "label": "Total Interest (E)", "formula": "=SUM(B26:B27)"},
        {"type": "total", "label": "F. EBT", "formula": "=B24-B28"},
        {"type": "formula", "label": "G. TAX (on positive EBT)", "value": data["tax_calc"]},
        {"type": "total", "label": "H. PAT", "formula": "=B29-B30"},
        {"type": "formula", "label": "I. Depreciation Added Back", "value": data["depreciation"]},
        {"type": "total", "label": "J. CASH ACCRUAL", "formula": "=B31+B32"},
        {"type": "formula", "label": "K. Repayment of Term Loan", "value": data["repayment_tl"]},
        {"type": "total", "label": "L. NET CASH AVAILABLE", "formula": "=B33-B34"},
    ]

    bs_rows = [
        {"type": "section", "label": "Assets"},
        {"type": "data", "label": "Gross Fixed Assets", "value": data["gross_fa"]},
        {"type": "data", "label": "Less: Accumulated Depreciation", "value": -data["accum_dep"]},
        {"type": "data", "label": "Inventory", "value": data["inventory"]},
        {"type": "data", "label": "Debtors", "value": data["debtors"]},
        {"type": "data", "label": "Cash & Bank", "value": data["cash_bank"]},
        {"type": "data", "label": "Other Current Assets", "value": data["other_ca"]},
        {"type": "data", "label": "Loans & Advances", "value": data["loans_adv"]},
        {"type": "total", "label": "Total Assets", "formula": "=SUM(B4:B10)"},
        {"type": "section", "label": "Liabilities"},
        {"type": "data", "label": "Capital", "value": data["capital"]},
        {"type": "data", "label": "Reserves & Surplus", "value": data["reserves"]},
        {"type": "data", "label": "Term Loan", "value": data["term_loan"]},
        {"type": "data", "label": "Bank CC / OD", "value": data["bank_cc"]},
        {"type": "data", "label": "Sundry Creditors", "value": data["sundry_creditors"]},
        {"type": "data", "label": "Other Current Liabilities", "value": data["other_cl"]},
        {"type": "data", "label": "Statutory Liabilities", "value": data["statutory_liab"]},
        {"type": "total", "label": "Total Liabilities", "formula": "=SUM(B12:B18)"},
        {"type": "formula", "label": "Difference (Assets - Liabilities)", "formula": "=B11-B19"},
    ]

    ratios_rows = [
        {"type": "section", "label": "Liquidity & Leverage"},
        {
            "type": "formula",
            "label": "Current Ratio = Current Assets / Current Liabilities",
            "value": data["current_ratio"],
            "number_format": INDIAN_PERCENT_FMT,
        },
        {
            "type": "formula",
            "label": "Debt Equity Ratio = Outside Liabilities / Tangible Net Worth",
            "value": data["debt_equity_ratio"],
            "number_format": INDIAN_PERCENT_FMT,
        },
        {
            "type": "formula",
            "label": "EBITDA Margin % = EBITDA / Total Income",
            "value": data["ebitda_margin"],
            "number_format": INDIAN_PERCENT_FMT,
        },
        {
            "type": "formula",
            "label": "Net Profit Margin % = PAT / Total Income",
            "value": data["net_profit_margin"],
            "number_format": INDIAN_PERCENT_FMT,
        },
    ]

    dscr_rows = [
        {"type": "section", "label": "Debt Service Coverage Ratio"},
        {"type": "formula", "label": "Numerator = PAT + Depreciation + Interest on Term Loan", "value": data["dscr_numerator"]},
        {"type": "formula", "label": "Denominator = Interest on Term Loan + Repayment of Term Loan", "value": data["dscr_denominator"]},
        {"type": "total", "label": "DSCR = Numerator / Denominator", "formula": "=IF(B5=0,0,B4/B5)", "number_format": INDIAN_PERCENT_FMT},
    ]

    validation_rows = [
        {"type": "section", "label": "Validation Checks"},
        {"type": "formula", "label": "Balance Sheet Tallies", "value": 1 if abs(data["total_assets"] - data["total_liabilities"]) < 1 else 0, "number_format": "0"},
        {"type": "formula", "label": "EBT = EBIT - Interest", "value": 1 if abs(data["ebt"] - (data["ebit"] - data["total_interest"])) < 1 else 0, "number_format": "0"},
        {"type": "formula", "label": "PAT = EBT - Tax", "value": 1 if abs(data["pat"] - (data["ebt"] - data["tax_calc"])) < 1 else 0, "number_format": "0"},
        {"type": "formula", "label": "Net Cash = Cash Accrual - TL Repayment", "value": 1 if abs(data["net_cash_available"] - (data["cash_accrual"] - data["repayment_tl"])) < 1 else 0, "number_format": "0"},
    ]

    render_sheet("Profit & Loss", "CMA Report - Profit & Loss Statement", pnl_rows)
    render_sheet("Balance Sheet", "CMA Report - Balance Sheet", bs_rows)
    render_sheet("Ratios", "CMA Report - Financial Ratios", ratios_rows, amount_header="Value")
    render_sheet("DSCR", "CMA Report - DSCR Analysis", dscr_rows, amount_header="Value")
    render_sheet("Validation", "CMA Report - Validation", validation_rows, amount_header="Status")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --- Header ---
header_left, header_right = st.columns([2.4, 1.4], vertical_alignment="center")
with header_left:
    st.markdown(
        """
        <div class='top-shell'>
            <div class='business-name'>Apex Capital Advisory LLP</div>
            <div class='business-address'>904, Trade Center, BKC Annex, Mumbai • Maharashtra • India</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
with header_right:
    st.markdown("### Dashboard Controls")
    party_name = st.text_input("Party Name", value="")
    constitution = st.selectbox("Constitution", ["Proprietorship", "Partnership", "Pvt Ltd"], index=0)

st.markdown("<div class='main-heading'>Actual Year Input (Locked for Projection)</div>", unsafe_allow_html=True)

main_col, summary_col = st.columns([3.5, 1.2], gap="large")

with main_col:
    # A. Profit & Loss Inputs
    st.markdown("<div class='section-title'>A. Profit &amp; Loss Inputs</div>", unsafe_allow_html=True)
    a1, a2 = st.columns(2, gap="medium")

    with a1:
        st.markdown("<div class='card'><div class='card-title'>Income</div>", unsafe_allow_html=True)
        domestic_sale = num("Domestic Sale", "domestic_sale")
        export_sale = num("Export Sale", "export_sale")
        other_income = num("Other Income", "other_income")
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='card'><div class='card-title'>Operating Expenses</div>", unsafe_allow_html=True)
        salary_wages = num("Salary & Wages", "salary_wages")
        power_fuel = num("Power & Fuel", "power_fuel")
        rent_exp = num("Rent Expenses", "rent_exp")
        printing_stationery = num("Printing & Stationery", "printing_stationery")
        depreciation = num("Depreciation", "depreciation")
        other_expenditure = num("Other Expenditure", "other_expenditure")
        st.markdown("</div>", unsafe_allow_html=True)

    with a2:
        st.markdown("<div class='card'><div class='card-title'>Direct Cost</div>", unsafe_allow_html=True)
        opening_stock = num("Opening Stock", "opening_stock")
        purchases = num("Purchases", "purchases")
        carriage_outward = num("Carriage Outward", "carriage_outward")
        unloading_expenses = num("Unloading Expenses", "unloading_expenses")
        direct_expenses = num("Direct Expenses", "direct_expenses")
        closing_stock = num("Closing Stock", "closing_stock")
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='card'><div class='card-title'>Finance & Tax</div>", unsafe_allow_html=True)
        interest_cc = num("Interest on CC", "interest_cc")
        interest_tl = num("Interest on TL", "interest_tl")
        tax_expense = num("Tax Expense", "tax_expense")
        repayment_tl = num("Repayment of Term Loan", "repayment_tl")
        st.markdown("</div>", unsafe_allow_html=True)

    # B. Balance Sheet Inputs
    st.markdown("<div class='section-title'>B. Balance Sheet Inputs</div>", unsafe_allow_html=True)
    b1, b2 = st.columns(2, gap="medium")
    with b1:
        st.markdown("<div class='card'><div class='card-title'>Assets</div>", unsafe_allow_html=True)
        gross_fa = num("Gross Fixed Assets", "gross_fa")
        accum_dep = num("Accumulated Depreciation", "accum_dep")
        inventory = num("Inventory", "inventory")
        debtors = num("Debtors", "debtors")
        cash_bank = num("Cash & Bank", "cash_bank")
        other_ca = num("Other Current Assets", "other_ca")
        loans_adv = num("Loans & Advances", "loans_adv")
        st.markdown("</div>", unsafe_allow_html=True)

    with b2:
        st.markdown("<div class='card'><div class='card-title'>Liabilities</div>", unsafe_allow_html=True)
        capital = num("Capital", "capital")
        reserves = num("Reserves & Surplus", "reserves")
        term_loan = num("Term Loan", "term_loan")
        bank_cc = num("Bank CC / OD", "bank_cc")
        sundry_creditors = num("Sundry Creditors", "sundry_creditors")
        other_cl = num("Other Current Liabilities", "other_cl")
        statutory_liab = num("Statutory Liabilities", "statutory_liab")
        st.markdown("</div>", unsafe_allow_html=True)

    # C. Projection Controls
    st.markdown("<div class='section-title'>C. Projection Controls</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3, gap="medium")
    with c1:
        st.markdown("<div class='card'><div class='card-title'>Growth</div>", unsafe_allow_html=True)
        domestic_growth = num("Domestic Sale Growth %", "domestic_growth")
        export_growth = num("Export Sale Growth %", "export_growth")
        other_income_growth = num("Other Income Growth %", "other_income_growth")
        material_growth = num("Material Cost Growth %", "material_growth")
        operating_growth = num("Operating Expense Growth %", "operating_growth")
        st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        st.markdown("<div class='card'><div class='card-title'>Working Capital Days</div>", unsafe_allow_html=True)
        debtor_days = num("Debtor Days", "debtor_days")
        creditor_days = num("Creditor Days", "creditor_days")
        inventory_days = num("Inventory Days", "inventory_days")
        st.markdown("</div>", unsafe_allow_html=True)

    with c3:
        st.markdown("<div class='card'><div class='card-title'>Rates</div>", unsafe_allow_html=True)
        cc_rate = num("Interest Rate on CC", "cc_rate")
        tl_rate = num("Interest Rate on TL", "tl_rate")
        tax_rate = num("Tax Rate", "tax_rate")
        st.markdown("</div>", unsafe_allow_html=True)

    # Buttons
    st.markdown("<div class='button-shell'>", unsafe_allow_html=True)
    btn1, btn2, btn3 = st.columns(3)
    with btn1:
        if st.button("Generate CMA", type="primary"):
            st.success(f"CMA prepared for {party_name or 'selected party'} ({constitution}).")
    with btn2:
        if st.button("Reset Inputs"):
            for k in list(st.session_state.keys()):
                if k not in ["party_name", "constitution"]:
                    st.session_state[k] = 0.0
            st.experimental_rerun()

    total_income = domestic_sale + export_sale + other_income
    material_cost = opening_stock + purchases + carriage_outward + unloading_expenses + direct_expenses - closing_stock
    operating_exp = salary_wages + power_fuel + rent_exp + printing_stationery + depreciation + other_expenditure
    ebit = total_income - material_cost - operating_exp
    total_interest = interest_cc + interest_tl
    ebt = ebit - total_interest
    tax_calc = max(0.0, ebt) * (tax_rate / 100.0) if tax_rate else tax_expense
    pat = ebt - tax_calc
    cash_accrual = pat + depreciation
    net_cash_available = cash_accrual - repayment_tl

    net_fixed_assets = gross_fa - accum_dep
    total_assets = net_fixed_assets + inventory + debtors + cash_bank + other_ca + loans_adv
    total_liabilities = capital + reserves + term_loan + bank_cc + sundry_creditors + other_cl + statutory_liab

    current_assets = inventory + debtors + cash_bank + other_ca + loans_adv
    current_liabilities = bank_cc + sundry_creditors + other_cl + statutory_liab
    outside_liabilities = term_loan + bank_cc + sundry_creditors + other_cl + statutory_liab
    tangible_net_worth = capital + reserves

    current_ratio = current_assets / current_liabilities if current_liabilities else 0.0
    debt_equity_ratio = outside_liabilities / tangible_net_worth if tangible_net_worth else 0.0
    ebitda = ebit + depreciation
    ebitda_margin = (ebitda / total_income) * 100 if total_income else 0.0
    net_profit_margin = (pat / total_income) * 100 if total_income else 0.0

    dscr_numerator = pat + depreciation + interest_tl
    dscr_denominator = interest_tl + repayment_tl
    dscr = dscr_numerator / dscr_denominator if dscr_denominator else 0.0

    export_payload = {
        "party_name": party_name,
        "constitution": constitution,
        "domestic_sale": domestic_sale,
        "export_sale": export_sale,
        "other_income": other_income,
        "total_income": total_income,
        "opening_stock": opening_stock,
        "purchases": purchases,
        "carriage_outward": carriage_outward,
        "unloading_expenses": unloading_expenses,
        "direct_expenses": direct_expenses,
        "closing_stock": closing_stock,
        "material_cost": material_cost,
        "salary_wages": salary_wages,
        "power_fuel": power_fuel,
        "rent_exp": rent_exp,
        "printing_stationery": printing_stationery,
        "depreciation": depreciation,
        "other_expenditure": other_expenditure,
        "operating_exp": operating_exp,
        "ebit": ebit,
        "interest_cc": interest_cc,
        "interest_tl": interest_tl,
        "total_interest": total_interest,
        "ebt": ebt,
        "tax_calc": tax_calc,
        "pat": pat,
        "cash_accrual": cash_accrual,
        "repayment_tl": repayment_tl,
        "net_cash_available": net_cash_available,
        "gross_fa": gross_fa,
        "accum_dep": accum_dep,
        "inventory": inventory,
        "debtors": debtors,
        "cash_bank": cash_bank,
        "other_ca": other_ca,
        "loans_adv": loans_adv,
        "capital": capital,
        "reserves": reserves,
        "term_loan": term_loan,
        "bank_cc": bank_cc,
        "sundry_creditors": sundry_creditors,
        "other_cl": other_cl,
        "statutory_liab": statutory_liab,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "current_ratio": current_ratio,
        "debt_equity_ratio": debt_equity_ratio,
        "ebitda_margin": ebitda_margin,
        "net_profit_margin": net_profit_margin,
        "dscr_numerator": dscr_numerator,
        "dscr_denominator": dscr_denominator,
        "dscr": dscr,
    }

    output = build_styled_cma_workbook(export_payload)
    with btn3:
        st.download_button(
            label="Export Excel",
            data=output,
            file_name="cma_dashboard_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    st.markdown("</div>", unsafe_allow_html=True)

# Summary calculations
material_cost = opening_stock + purchases + carriage_outward + unloading_expenses + direct_expenses - closing_stock
operating_exp = salary_wages + power_fuel + rent_exp + printing_stationery + depreciation + other_expenditure
total_income = domestic_sale + export_sale + other_income
ebit = total_income - material_cost - operating_exp
total_interest = interest_cc + interest_tl
ebt = ebit - total_interest
tax_calc = max(0.0, ebt) * (tax_rate / 100.0) if tax_rate else tax_expense
pat = ebt - tax_calc
cash_accrual = pat + depreciation
working_capital = inventory + debtors + other_ca + loans_adv - sundry_creditors - other_cl - statutory_liab

summary = Summary(
    total_income=total_income,
    material_cost=material_cost,
    ebit=ebit,
    total_interest=total_interest,
    ebt=ebt,
    pat=pat,
    cash_accrual=cash_accrual,
    working_capital=working_capital,
)

with summary_col:
    st.markdown("<div class='summary-box'><div class='summary-title'>Calculated Summary</div>", unsafe_allow_html=True)

    def metric(label: str, value: float):
        st.markdown(
            f"<div class='metric-row'><span class='metric-label'>{label}</span><span class='metric-val'>{value:,.2f}</span></div>",
            unsafe_allow_html=True,
        )

    metric("Total Income", summary.total_income)
    metric("Material Cost", summary.material_cost)
    metric("EBIT", summary.ebit)
    metric("Total Interest", summary.total_interest)
    metric("EBT", summary.ebt)
    metric("PAT", summary.pat)
    metric("Cash Accrual", summary.cash_accrual)
    metric("Working Capital", summary.working_capital)

    st.markdown("</div>", unsafe_allow_html=True)
