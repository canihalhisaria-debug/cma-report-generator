from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


YEARS = ["FY 1", "FY 2", "FY 3", "FY 4", "FY 5"]

HEADER_FILL = PatternFill("solid", fgColor="103C72")
SECTION_FILL = PatternFill("solid", fgColor="CFDCC8")
SUBSECTION_FILL = PatternFill("solid", fgColor="C7D1DE")
TOTAL_FILL = PatternFill("solid", fgColor="EAD8C5")
THIN = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)


def safe_div(a: float, b: float) -> float:
    return 0.0 if b == 0 else a / b


def growth_series(base: float, growth_rates: list[float]) -> list[float]:
    vals = [base]
    for rate in growth_rates[1:]:
        vals.append(vals[-1] * (1 + rate))
    return vals


def rolling_opening(closing_values: list[float]) -> list[float]:
    return [0.0] + closing_values[:-1]


def fmt_no_round(v: float) -> float:
    return round(v, 2)


def write_table(
    ws,
    title: str,
    rows: list[tuple[str, list[float] | list[str] | None, str | None, str]],
) -> None:
    ws.append([title] + YEARS)
    for c in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    for label, vals, style, num_fmt in rows:
        ws.append([label] + (vals if vals is not None else [""] * 5))
        r = ws.max_row
        fill = None
        if style == "section":
            fill = SECTION_FILL
        elif style == "subsection":
            fill = SUBSECTION_FILL
        elif style == "total":
            fill = TOTAL_FILL

        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
                if style in {"section", "subsection", "total"} or vals is None:
                    cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 48
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16


def build_new_cc_projection(cc_amount: float, roi: float) -> dict[str, Any]:
    years = 5

    # -------------------------------
    # REVERSE LOGIC ASSUMPTIONS
    # -------------------------------
    # Target: MPBF should be equal to or slightly above requested CC,
    # sales should be logical, ratios should remain banker-friendly.
    cogs_pct = 0.69
    opex_pct = 0.19
    tax_rate = 0.25

    debtor_days = 90
    inventory_days = 70
    creditor_days = 43

    cash_pct = 0.02
    other_ca_pct = 0.01
    statutory_pct = 0.015
    other_cl_pct = 0.004

    drawings_pct = 0.80  # keeps surplus controlled so MPBF does not become too high
    growth_rates = [0.0, 0.12, 0.10, 0.08, 0.07]

    # Reverse logic: compute FY1 sales from requested CC
    mpbf_coeff = (
        0.75
        * (
            debtor_days / 365
            + (cogs_pct * inventory_days / 365)
            + cash_pct
            + other_ca_pct
        )
        - (cogs_pct * creditor_days / 365 + statutory_pct + other_cl_pct)
    )

    # Keep WCA slightly above requested CC, not too much
    target_mpbf_fy1 = cc_amount * 1.05
    fy1_sales = target_mpbf_fy1 / mpbf_coeff
    sales = [fmt_no_round(v) for v in growth_series(fy1_sales, growth_rates)]

    # -------------------------------
    # STOCK / PURCHASE / RECEIVABLES
    # -------------------------------
    closing_stock = [fmt_no_round(s * cogs_pct * inventory_days / 365) for s in sales]
    opening_stock = [fmt_no_round(v) for v in rolling_opening(closing_stock)]
    debtors = [fmt_no_round(s * debtor_days / 365) for s in sales]
    creditors = [fmt_no_round(s * cogs_pct * creditor_days / 365) for s in sales]
    cash_bank = [fmt_no_round(s * cash_pct) for s in sales]
    other_current_assets_base = [fmt_no_round(s * other_ca_pct) for s in sales]
    statutory_liabilities = [fmt_no_round(s * statutory_pct) for s in sales]
    other_current_liabilities = [fmt_no_round(s * other_cl_pct) for s in sales]

    cost_of_sales = [fmt_no_round(s * cogs_pct) for s in sales]
    purchases = [
        fmt_no_round(cogs + cs - os)
        for cogs, cs, os in zip(cost_of_sales, closing_stock, opening_stock)
    ]
    carriage_inward = [0.0] * years

    # -------------------------------
    # FIXED ASSETS / DEPRECIATION
    # -------------------------------
    initial_gross_block = max(cc_amount * 0.28, sales[0] * 0.10)

    dep_rate = 0.10
    opening_gross_block: list[float] = []
    additions: list[float] = []
    gross_block: list[float] = []
    opening_acc_dep: list[float] = []
    depreciation: list[float] = []
    accumulated_dep: list[float] = []
    net_block: list[float] = []

    prev_gross = initial_gross_block
    prev_acc = initial_gross_block * 0.08

    for i in range(years):
        add = 0.0 if i == 0 else max((sales[i] - sales[i - 1]) * 0.04, 0.0)
        dep = (prev_gross * dep_rate) + (add * dep_rate * 0.5)
        curr_gross = prev_gross + add
        curr_acc = prev_acc + dep
        curr_net = max(curr_gross - curr_acc, 0.0)

        opening_gross_block.append(fmt_no_round(prev_gross))
        additions.append(fmt_no_round(add))
        gross_block.append(fmt_no_round(curr_gross))
        opening_acc_dep.append(fmt_no_round(prev_acc))
        depreciation.append(fmt_no_round(dep))
        accumulated_dep.append(fmt_no_round(curr_acc))
        net_block.append(fmt_no_round(curr_net))

        prev_gross = curr_gross
        prev_acc = curr_acc

    # -------------------------------
    # P&L
    # -------------------------------
    domestic_sales = sales[:]
    export_sales = [0.0] * years
    other_income = [0.0] * years

    salary = [fmt_no_round(s * 0.055) for s in sales]
    rent = [fmt_no_round(s * 0.020) for s in sales]
    power_fuel = [fmt_no_round(s * 0.018) for s in sales]
    travelling = [fmt_no_round(s * 0.008) for s in sales]
    telephone = [fmt_no_round(s * 0.004) for s in sales]
    office = [fmt_no_round(s * 0.010) for s in sales]
    printing = [fmt_no_round(s * 0.003) for s in sales]
    repairs = [fmt_no_round(s * 0.006) for s in sales]

    operating_expenses = [fmt_no_round(s * opex_pct) for s in sales]
    other_operating = [
        fmt_no_round(
            opex
            - (sa + re + po + tr + te + of + pr + rp)
        )
        for opex, sa, re, po, tr, te, of, pr, rp in zip(
            operating_expenses,
            salary,
            rent,
            power_fuel,
            travelling,
            telephone,
            office,
            printing,
            repairs,
        )
    ]

    op_profit_before_int_dep = [
        fmt_no_round(s - c - opex)
        for s, c, opex in zip(sales, cost_of_sales, operating_expenses)
    ]

    interest_cc = [fmt_no_round(cc_amount * roi / 100)] * years
    op_profit_after_int_dep = [
        fmt_no_round(p - i - d)
        for p, i, d in zip(op_profit_before_int_dep, interest_cc, depreciation)
    ]

    pbt = [fmt_no_round(v + oi) for v, oi in zip(op_profit_after_int_dep, other_income)]
    tax = [fmt_no_round(max(v * tax_rate, 0.0)) for v in pbt]
    net_profit = [fmt_no_round(v - t) for v, t in zip(pbt, tax)]
    drawings = [fmt_no_round(np * drawings_pct) for np in net_profit]
    retained_profit = [fmt_no_round(np - dr) for np, dr in zip(net_profit, drawings)]

    # -------------------------------
    # BALANCE SHEET LOGIC
    # -------------------------------
    required_nw: list[float] = []
    cumulative_surplus: list[float] = []
    run_surplus = 0.0

    for i in range(years):
        base_ca = debtors[i] + closing_stock[i] + cash_bank[i] + other_current_assets_base[i]
        outside_liab_without_cc_tax = (
            creditors[i] + statutory_liabilities[i] + other_current_liabilities[i]
        )
        outside_liab = cc_amount + outside_liab_without_cc_tax + tax[i]
        req_nw = base_ca + net_block[i] - outside_liab
        required_nw.append(fmt_no_round(req_nw))

        run_surplus += retained_profit[i]
        cumulative_surplus.append(fmt_no_round(run_surplus))

    share_capital_value = max(
        max(req - sur for req, sur in zip(required_nw, cumulative_surplus)),
        cc_amount * 0.12,
    )
    share_capital = [fmt_no_round(share_capital_value)] * years
    surplus_pl = cumulative_surplus[:]

    other_current_assets: list[float] = []
    total_current_assets: list[float] = []
    total_current_liabilities: list[float] = []
    total_outside_liabilities: list[float] = []
    total_assets: list[float] = []
    total_liabilities: list[float] = []
    net_worth: list[float] = []
    balance_diff: list[float] = []
    balance_status: list[str] = []

    for i in range(years):
        nw = fmt_no_round(share_capital[i] + surplus_pl[i])
        extra_oca = max(nw - required_nw[i], 0.0)

        oca = fmt_no_round(other_current_assets_base[i] + extra_oca)
        ca_total = fmt_no_round(debtors[i] + closing_stock[i] + cash_bank[i] + oca)
        cl_total = fmt_no_round(
            cc_amount
            + creditors[i]
            + statutory_liabilities[i]
            + other_current_liabilities[i]
            + tax[i]
        )
        outside = fmt_no_round(cl_total)
        assets = fmt_no_round(ca_total + net_block[i])
        liabs = fmt_no_round(outside + nw)

        other_current_assets.append(oca)
        total_current_assets.append(ca_total)
        total_current_liabilities.append(cl_total)
        total_outside_liabilities.append(outside)
        total_assets.append(assets)
        total_liabilities.append(liabs)
        net_worth.append(nw)

        diff = fmt_no_round(assets - liabs)
        balance_diff.append(diff)
        balance_status.append("OK" if abs(diff) <= 1 else "CHECK")

    net_working_capital = [
        fmt_no_round(a - b)
        for a, b in zip(total_current_assets, total_current_liabilities)
    ]

    # -------------------------------
    # WORKING CAPITAL ANALYSIS
    # -------------------------------
    wc_total_cl_other = [
        fmt_no_round(c + s + o + t)
        for c, s, o, t in zip(
            creditors,
            statutory_liabilities,
            other_current_liabilities,
            tax,
        )
    ]
    wc_gap = [
        fmt_no_round(a - b)
        for a, b in zip(total_current_assets, wc_total_cl_other)
    ]
    borrower_contribution = [fmt_no_round(a * 0.25) for a in total_current_assets]
    mpbf = [
        fmt_no_round(max((a * 0.75) - b, 0.0))
        for a, b in zip(total_current_assets, wc_total_cl_other)
    ]
    proposed_cc_limit = [fmt_no_round(cc_amount)] * years

    nayak_wc_requirement = [fmt_no_round(s * 0.25) for s in sales]
    nayak_borrower_contribution = [fmt_no_round(s * 0.05) for s in sales]
    nayak_eligible_bank_finance = [fmt_no_round(s * 0.20) for s in sales]

    # -------------------------------
    # RATIOS
    # -------------------------------
    def bench(name: str, value: float) -> tuple[str, str]:
        rules = {
            "Current Ratio": (1.33, 2.50),
            "Quick Ratio": (1.00, 2.00),
            "Debt Equity Ratio": (0.00, 2.00),
            "Debt Asset Ratio": (0.00, 0.75),
            "Gross Profit Ratio": (0.15, 0.45),
            "Operating Profit Ratio": (0.08, 0.25),
            "Net Profit Ratio": (0.05, 0.20),
            "Interest Coverage Ratio": (1.50, 99.00),
            "DSCR": (1.25, 99.00),
            "Inventory Turnover": (4.00, 12.00),
            "Debtors Turnover": (3.00, 10.00),
            "Creditors Turnover": (4.00, 12.00),
            "Working Capital Turnover": (3.00, 8.00),
        }
        lo, hi = rules[name]
        return (f"{lo:.2f} - {hi:.2f}", "OK" if lo <= value <= hi else "Alert")

    ratios: list[dict[str, Any]] = []

    def add_ratio(
        sno: int,
        name: str,
        num_label: str,
        den_label: str,
        values: list[float],
    ) -> None:
        bm, st = bench(name, values[0])
        ratios.append(
            {
                "S.No": sno,
                "Particulars": name,
                "Numerator": num_label,
                "Denominator": den_label,
                "FY1": fmt_no_round(values[0]),
                "FY2": fmt_no_round(values[1]),
                "FY3": fmt_no_round(values[2]),
                "FY4": fmt_no_round(values[3]),
                "FY5": fmt_no_round(values[4]),
                "Bank Acceptable Benchmark": bm,
                "Status": st,
            }
        )

    add_ratio(
        1,
        "Current Ratio",
        "Current Assets",
        "Current Liabilities",
        [safe_div(a, b) for a, b in zip(total_current_assets, total_current_liabilities)],
    )
    add_ratio(
        2,
        "Quick Ratio",
        "Current Assets - Inventory",
        "Current Liabilities",
        [
            safe_div(a - inv, b)
            for a, inv, b in zip(
                total_current_assets,
                closing_stock,
                total_current_liabilities,
            )
        ],
    )
    add_ratio(
        3,
        "Debt Equity Ratio",
        "Total Debt",
        "Net Worth",
        [safe_div(d, nw) for d, nw in zip(total_outside_liabilities, net_worth)],
    )
    add_ratio(
        4,
        "Debt Asset Ratio",
        "Total Debt",
        "Total Assets",
        [safe_div(d, a) for d, a in zip(total_outside_liabilities, total_assets)],
    )
    add_ratio(
        5,
        "Gross Profit Ratio",
        "Gross Profit",
        "Net Sales",
        [safe_div(s - c, s) for s, c in zip(sales, cost_of_sales)],
    )
    add_ratio(
        6,
        "Operating Profit Ratio",
        "Operating Profit",
        "Net Sales",
        [safe_div(op, s) for op, s in zip(op_profit_before_int_dep, sales)],
    )
    add_ratio(
        7,
        "Net Profit Ratio",
        "Net Profit",
        "Net Sales",
        [safe_div(np, s) for np, s in zip(net_profit, sales)],
    )
    add_ratio(
        8,
        "Interest Coverage Ratio",
        "EBIT",
        "Interest",
        [safe_div(p + i, i) for p, i in zip(op_profit_after_int_dep, interest_cc)],
    )
    add_ratio(
        9,
        "DSCR",
        "Cash Profit",
        "Debt Service",
        [safe_div(np + dep, i) for np, dep, i in zip(net_profit, depreciation, interest_cc)],
    )
    add_ratio(
        10,
        "Inventory Turnover",
        "COGS",
        "Inventory",
        [safe_div(c, inv) for c, inv in zip(cost_of_sales, closing_stock)],
    )
    add_ratio(
        11,
        "Debtors Turnover",
        "Sales",
        "Debtors",
        [safe_div(s, d) for s, d in zip(sales, debtors)],
    )
    add_ratio(
        12,
        "Creditors Turnover",
        "Purchases",
        "Creditors",
        [safe_div(p, c) for p, c in zip(purchases, creditors)],
    )
    add_ratio(
        13,
        "Working Capital Turnover",
        "Sales",
        "Working Capital",
        [safe_div(s, wc) for s, wc in zip(sales, net_working_capital)],
    )

    assumptions = {
        "business_type": "New Business",
        "loan_type": "New CC",
        "cc_amount": fmt_no_round(cc_amount),
        "roi_pct": fmt_no_round(roi),
        "cogs_pct": fmt_no_round(cogs_pct * 100),
        "opex_pct": fmt_no_round(opex_pct * 100),
        "debtor_days": debtor_days,
        "inventory_days": inventory_days,
        "creditor_days": creditor_days,
        "drawings_pct": fmt_no_round(drawings_pct * 100),
        "sales_reverse_logic_fy1": fmt_no_round(fy1_sales),
        "mpbf_target_fy1": fmt_no_round(target_mpbf_fy1),
    }

    return {
        "assumptions": assumptions,
        "pl": {
            "sales": sales,
            "domestic_sales": domestic_sales,
            "export_sales": export_sales,
            "sales_growth_pct": [
                0.0,
                fmt_no_round((sales[1] - sales[0]) / sales[0] * 100),
                fmt_no_round((sales[2] - sales[1]) / sales[1] * 100),
                fmt_no_round((sales[3] - sales[2]) / sales[2] * 100),
                fmt_no_round((sales[4] - sales[3]) / sales[3] * 100),
            ],
            "other_income": other_income,
            "purchases": purchases,
            "carriage_inward": carriage_inward,
            "opening_stock": opening_stock,
            "closing_stock": closing_stock,
            "cost_of_sales": cost_of_sales,
            "salary": salary,
            "rent": rent,
            "power_fuel": power_fuel,
            "travelling": travelling,
            "telephone": telephone,
            "office": office,
            "printing": printing,
            "repairs": repairs,
            "other_operating": other_operating,
            "operating_expenses": operating_expenses,
            "op_profit_before_int_dep": op_profit_before_int_dep,
            "interest_cc": interest_cc,
            "depreciation": depreciation,
            "op_profit_after_int_dep": op_profit_after_int_dep,
            "pbt": pbt,
            "tax": tax,
            "net_profit": net_profit,
            "drawings": drawings,
            "retained_profit": retained_profit,
            "dep_opening_gross_block": opening_gross_block,
            "dep_additions": additions,
            "dep_gross_block": gross_block,
            "dep_opening_accumulated": opening_acc_dep,
            "dep_accumulated": accumulated_dep,
            "dep_net_block": net_block,
        },
        "bs": {
            "cc_from_applicant_bank": [fmt_no_round(cc_amount)] * years,
            "cc_from_other_banks": [0.0] * years,
            "bills_purchased": [0.0] * years,
            "short_term_borrowings_others": [0.0] * years,
            "sundry_creditors": creditors,
            "advances_customers": [0.0] * years,
            "provision_tax": tax,
            "dividend_payable": [0.0] * years,
            "other_statutory": statutory_liabilities,
            "other_current_liabilities": other_current_liabilities,
            "total_current_liabilities": total_current_liabilities,
            "term_loans": [0.0] * years,
            "other_term_liabilities": [0.0] * years,
            "total_outside_liabilities": total_outside_liabilities,
            "share_capital": share_capital,
            "general_reserve": [0.0] * years,
            "revaluation_reserve": [0.0] * years,
            "other_reserves": [0.0] * years,
            "surplus_pl": surplus_pl,
            "net_worth": net_worth,
            "total_liabilities": total_liabilities,
            "cash_bank": cash_bank,
            "govt_securities": [0.0] * years,
            "fixed_deposits": [0.0] * years,
            "receivables": debtors,
            "export_receivables": [0.0] * years,
            "deferred_receivables": [0.0] * years,
            "stocks": closing_stock,
            "advances_suppliers": [0.0] * years,
            "advance_tax": [0.0] * years,
            "other_current_assets": other_current_assets,
            "total_current_assets": total_current_assets,
            "gross_block": gross_block,
            "dep_to_date": accumulated_dep,
            "net_block": net_block,
            "other_investments": [0.0] * years,
            "security_deposits": [0.0] * years,
            "other_non_current_assets": [0.0] * years,
            "total_other_non_current": [0.0] * years,
            "intangible_assets": [0.0] * years,
            "total_assets": total_assets,
            "net_working_capital": net_working_capital,
            "balance_diff": balance_diff,
            "balance_status": balance_status,
        },
        "wc": {
            "raw_material": [0.0] * years,
            "wip": [0.0] * years,
            "finished_goods": closing_stock,
            "receivables": debtors,
            "cash_bank": cash_bank,
            "other_current_assets": other_current_assets,
            "total_current_assets": total_current_assets,
            "sundry_creditors": creditors,
            "outstanding_expenses": other_current_liabilities,
            "statutory_liabilities": [
                fmt_no_round(s + t) for s, t in zip(statutory_liabilities, tax)
            ],
            "total_current_liabilities": wc_total_cl_other,
            "wc_gap": wc_gap,
            "borrower_contribution": borrower_contribution,
            "mpbf": mpbf,
            "proposed_cc_limit": proposed_cc_limit,
            "projected_annual_turnover": sales,
            "nayak_wc_requirement": nayak_wc_requirement,
            "nayak_borrower_contribution": nayak_borrower_contribution,
            "nayak_eligible_bank_finance": nayak_eligible_bank_finance,
        },
        "ratios": ratios,
    }


def build_workbook(output_path: Path, result: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    assumptions = result["assumptions"]
    pl = result["pl"]
    bs = result["bs"]
    wc = result["wc"]
    ratios = result["ratios"]

    ws_a = wb.create_sheet("ASSUMPTIONS")
    ws_a.append(["Parameter", "Value"])
    for row in assumptions.items():
        ws_a.append([row[0], row[1]])
    for cell in ws_a[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN
    for r in ws_a.iter_rows(min_row=2):
        for c in r:
            c.border = THIN
    ws_a.column_dimensions["A"].width = 28
    ws_a.column_dimensions["B"].width = 18

    write_table(
        wb.create_sheet("PROJECTED PL"),
        "PROJECTED PL",
        [
            ("1. Gross Income", None, "section", "#,##0.00"),
            ("(i) Sales (Net of Returns)", None, "subsection", "#,##0.00"),
            ("(a) Domestic Sales", pl["domestic_sales"], None, "#,##0.00"),
            ("(b) Export Sales", pl["export_sales"], None, "#,##0.00"),
            ("(c) Sub-Total", pl["sales"], "total", "#,##0.00"),
            ("(d) Percentage Rise/Fall in Sales", pl["sales_growth_pct"], None, "0.00"),
            ("(ii) Other Income", None, "subsection", "#,##0.00"),
            ("(a) Duty Drawback", [0.0] * 5, None, "#,##0.00"),
            ("(b) Cash Assistance", [0.0] * 5, None, "#,##0.00"),
            ("(c) Commission & Brokerage", [0.0] * 5, None, "#,##0.00"),
            ("(d) Sub-Total", pl["other_income"], "total", "#,##0.00"),
            (
                "(iii) Total Gross Income",
                [a + b for a, b in zip(pl["sales"], pl["other_income"])],
                "section",
                "#,##0.00",
            ),
            ("2. Cost of Sales", None, "section", "#,##0.00"),
            ("(i) Purchases", pl["purchases"], None, "#,##0.00"),
            ("(ii) Carriage Inward", pl["carriage_inward"], None, "#,##0.00"),
            (
                "(iii) Sub-Total",
                [a + b for a, b in zip(pl["purchases"], pl["carriage_inward"])],
                "total",
                "#,##0.00",
            ),
            ("(iv) Add Opening Stock", pl["opening_stock"], None, "#,##0.00"),
            ("(vi) Less Closing Stock", pl["closing_stock"], None, "#,##0.00"),
            ("(vii) Total Cost of Sales", pl["cost_of_sales"], "section", "#,##0.00"),
            ("3. Operating Expenses", None, "section", "#,##0.00"),
            ("Salary", pl["salary"], None, "#,##0.00"),
            ("Rent", pl["rent"], None, "#,##0.00"),
            ("Power & Fuel", pl["power_fuel"], None, "#,##0.00"),
            ("Travelling & Conveyance", pl["travelling"], None, "#,##0.00"),
            ("Telephone & Internet", pl["telephone"], None, "#,##0.00"),
            ("Office Expenses", pl["office"], None, "#,##0.00"),
            ("Printing & Stationery", pl["printing"], None, "#,##0.00"),
            ("Repairs & Maintenance", pl["repairs"], None, "#,##0.00"),
            ("Other Operating Expenses", pl["other_operating"], None, "#,##0.00"),
            (
                "4. Operating Profit (Before Interest & Depreciation)",
                pl["op_profit_before_int_dep"],
                "section",
                "#,##0.00",
            ),
            ("5. Interest on Cash Credit / OD", pl["interest_cc"], "section", "#,##0.00"),
            ("6. Depreciation", pl["depreciation"], "section", "#,##0.00"),
            (
                "7. Operating Profit (After Interest & Depreciation)",
                pl["op_profit_after_int_dep"],
                "section",
                "#,##0.00",
            ),
            ("9. Profit Before Tax", pl["pbt"], "section", "#,##0.00"),
            ("10. Provision for Tax", pl["tax"], "subsection", "#,##0.00"),
            ("11. Net Profit", pl["net_profit"], "section", "#,##0.00"),
            ("12. Drawings", pl["drawings"], "subsection", "#,##0.00"),
            ("13. Retained Profit", pl["retained_profit"], "total", "#,##0.00"),
        ],
    )

    write_table(
        wb.create_sheet("PROJECTED BS"),
        "PROJECTED BS",
        [
            ("CURRENT LIABILITIES", None, "section", "#,##0.00"),
            (
                "Short Term Borrowings from Applicant Bank",
                bs["cc_from_applicant_bank"],
                None,
                "#,##0.00",
            ),
            ("Sundry Creditors (Trade)", bs["sundry_creditors"], None, "#,##0.00"),
            ("Provision for Taxation", bs["provision_tax"], None, "#,##0.00"),
            (
                "Other Statutory Liabilities",
                bs["other_statutory"],
                None,
                "#,##0.00",
            ),
            (
                "Other Current Liabilities",
                bs["other_current_liabilities"],
                None,
                "#,##0.00",
            ),
            (
                "Total Current Liabilities (B)",
                bs["total_current_liabilities"],
                "total",
                "#,##0.00",
            ),
            ("TERM LIABILITIES", None, "section", "#,##0.00"),
            ("Term Loans", bs["term_loans"], None, "#,##0.00"),
            (
                "Total Outside Liabilities (D)",
                bs["total_outside_liabilities"],
                "section",
                "#,##0.00",
            ),
            ("NET WORTH", None, "section", "#,##0.00"),
            ("Share Capital", bs["share_capital"], None, "#,##0.00"),
            (
                "Surplus / (Deficit) in P&L A/c",
                bs["surplus_pl"],
                None,
                "#,##0.00",
            ),
            ("Net Worth (E)", bs["net_worth"], "total", "#,##0.00"),
            ("Total Liabilities (F)", bs["total_liabilities"], "section", "#,##0.00"),
            ("CURRENT ASSETS", None, "section", "#,##0.00"),
            ("Cash & Bank Balances", bs["cash_bank"], None, "#,##0.00"),
            ("Receivables", bs["receivables"], None, "#,##0.00"),
            ("Stocks-in-Trade", bs["stocks"], None, "#,##0.00"),
            (
                "Other Current Assets",
                bs["other_current_assets"],
                None,
                "#,##0.00",
            ),
            (
                "Total Current Assets (G)",
                bs["total_current_assets"],
                "total",
                "#,##0.00",
            ),
            ("FIXED ASSETS", None, "section", "#,##0.00"),
            ("Gross Block", bs["gross_block"], None, "#,##0.00"),
            ("Depreciation to Date", bs["dep_to_date"], None, "#,##0.00"),
            ("Net Block (H)", bs["net_block"], "total", "#,##0.00"),
            ("Total Assets (J)", bs["total_assets"], "section", "#,##0.00"),
            ("Net Working Capital (L)", bs["net_working_capital"], "total", "#,##0.00"),
            ("Balance Status", bs["balance_status"], "subsection", "@"),
        ],
    )

    write_table(
        wb.create_sheet("WORKING CAPITAL ANALYSIS"),
        "WORKING CAPITAL ANALYSIS",
        [
            ("A. CURRENT ASSETS", None, "section", "#,##0.00"),
            ("Raw Material", wc["raw_material"], None, "#,##0.00"),
            ("Work in Progress", wc["wip"], None, "#,##0.00"),
            ("Finished Goods", wc["finished_goods"], None, "#,##0.00"),
            ("Receivables", wc["receivables"], None, "#,##0.00"),
            ("Cash & Bank", wc["cash_bank"], None, "#,##0.00"),
            (
                "Other Current Assets",
                wc["other_current_assets"],
                None,
                "#,##0.00",
            ),
            ("Total Current Assets (A)", wc["total_current_assets"], "total", "#,##0.00"),
            (
                "B. CURRENT LIABILITIES (OTHER THAN BANK)",
                None,
                "section",
                "#,##0.00",
            ),
            ("Sundry Creditors", wc["sundry_creditors"], None, "#,##0.00"),
            ("Outstanding Expenses", wc["outstanding_expenses"], None, "#,##0.00"),
            (
                "Statutory Liabilities",
                wc["statutory_liabilities"],
                None,
                "#,##0.00",
            ),
            (
                "Total Current Liabilities (B)",
                wc["total_current_liabilities"],
                "total",
                "#,##0.00",
            ),
            ("C. Working Capital Gap (A - B)", wc["wc_gap"], "section", "#,##0.00"),
            (
                "D. Borrower Contribution (25% of CA)",
                wc["borrower_contribution"],
                "subsection",
                "#,##0.00",
            ),
            (
                "E. Maximum Permissible Bank Finance (MPBF)",
                wc["mpbf"],
                "total",
                "#,##0.00",
            ),
            ("F. Proposed CC Limit", wc["proposed_cc_limit"], "section", "#,##0.00"),
            (
                "Alternative - Projected Annual Turnover",
                wc["projected_annual_turnover"],
                "subsection",
                "#,##0.00",
            ),
            (
                "Working Capital Requirement @25%",
                wc["nayak_wc_requirement"],
                None,
                "#,##0.00",
            ),
            (
                "Borrower Contribution @5%",
                wc["nayak_borrower_contribution"],
                None,
                "#,##0.00",
            ),
            (
                "Eligible Bank Finance @20%",
                wc["nayak_eligible_bank_finance"],
                "total",
                "#,##0.00",
            ),
        ],
    )

    write_table(
        wb.create_sheet("DEPRECIATION SCHEDULE"),
        "DEPRECIATION SCHEDULE",
        [
            ("Opening Gross Block", pl["dep_opening_gross_block"], None, "#,##0.00"),
            ("Additions During Year", pl["dep_additions"], None, "#,##0.00"),
            ("Gross Block", pl["dep_gross_block"], "total", "#,##0.00"),
            (
                "Opening Accumulated Depreciation",
                pl["dep_opening_accumulated"],
                None,
                "#,##0.00",
            ),
            ("Depreciation for the Year", pl["depreciation"], "section", "#,##0.00"),
            ("Accumulated Depreciation", pl["dep_accumulated"], "total", "#,##0.00"),
            ("Closing Net Block", pl["dep_net_block"], "section", "#,##0.00"),
        ],
    )

    ws_r = wb.create_sheet("FINANCIAL RATIOS")
    ws_r.append(
        [
            "S.No",
            "Particulars",
            "Numerator",
            "Denominator",
            *YEARS,
            "Bank Acceptable Benchmark",
            "Status",
        ]
    )
    for cell in ws_r[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    for row in ratios:
        ws_r.append(
            [
                row["S.No"],
                row["Particulars"],
                row["Numerator"],
                row["Denominator"],
                row["FY1"],
                row["FY2"],
                row["FY3"],
                row["FY4"],
                row["FY5"],
                row["Bank Acceptable Benchmark"],
                row["Status"],
            ]
        )
    for r in ws_r.iter_rows(min_row=2):
        for c in r:
            c.border = THIN
            if isinstance(c.value, (int, float)) and c.column in {5, 6, 7, 8, 9}:
                c.number_format = "0.00"
    ws_r.column_dimensions["B"].width = 30
    ws_r.column_dimensions["C"].width = 22
    ws_r.column_dimensions["D"].width = 22
    ws_r.column_dimensions["J"].width = 22
    ws_r.column_dimensions["K"].width = 12

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cc-amount", type=float, required=True, help="Requested CC limit")
    parser.add_argument("--roi", type=float, required=True, help="ROI in %")
    parser.add_argument("--output", default="NEW_BUSINESS_NEW_CC_CMA.xlsx")
    args = parser.parse_args()

    result = build_new_cc_projection(args.cc_amount, args.roi)
    build_workbook(Path(args.output), result)

    print(f"CMA file created: {args.output}")
    print("Reverse logic used: sales derived from requested CC so that MPBF stays near requested CC.")
    print("Fixed assets and depreciation were auto-decided.")
    print("Projected PL, BS, WCA, ratios, and depreciation schedule generated.")


if __name__ == "__main__":
    main()
