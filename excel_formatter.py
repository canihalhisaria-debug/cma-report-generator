from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from formulas import YEARS


HEADER_FILL = PatternFill("solid", fgColor="103C72")
SECTION_FILL = PatternFill("solid", fgColor="CFDCC8")
SUBSECTION_FILL = PatternFill("solid", fgColor="C7D1DE")
TOTAL_FILL = PatternFill("solid", fgColor="EAD8C5")
DEFAULT_HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
THIN = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)


def _resolve_row_fill(style: str | None, values: list[float] | None):
    if style == "section":
        return SECTION_FILL
    if style == "subsection":
        return SUBSECTION_FILL
    if style == "total":
        return TOTAL_FILL
    if values is None:
        return SECTION_FILL
    return None


def _write_table(
    ws,
    title: str,
    rows: list[tuple[str, list[float] | None, str | None]],
    percent: bool = False,
) -> None:
    ws.append([title] + YEARS)
    for c in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    for label, vals, style in rows:
        ws.append([label] + (vals if vals is not None else [""] * 5))
        r = ws.max_row
        fill = _resolve_row_fill(style, vals)
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
                if vals is None or style in {"section", "subsection", "total"}:
                    cell.font = Font(bold=True)
            else:
                cell.number_format = "0.00%" if percent else "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill
                if c == 1:
                    cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 52
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16


def build_excel_report(result: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    pl = result["pl"]
    bs = result["bs"]
    wc = result["wc"]
    ratios = result["ratios"]

    ws_pl = wb.create_sheet("PROJECTED PL")
    pl_rows = [
        ("1. Gross Income", None, "section"),
        ("(i) Sales (Net of Returns)", None, "subsection"),
        ("(a) Domestic Sales", pl["domestic_sales"], None),
        ("(b) Export Sales", pl["export_sales"], None),
        ("(c) Sub-Total", pl["sales"], "total"),
        ("(d) Percentage Rise/Fall in Sales", pl["sales_growth_pct"], None),
        ("(ii) Other Income", None, "subsection"),
        ("(a) Duty Drawback", [0.0] * 5, None),
        ("(b) Cash Assistance", [0.0] * 5, None),
        ("(c) Commission & Brokerage", [0.0] * 5, None),
        ("(d) Sub-Total", pl["other_income"], "total"),
        ("(iii) Total Gross Income", [s + oi for s, oi in zip(pl["sales"], pl["other_income"])], "section"),
        ("2. Cost of Sales", None, "section"),
        ("(i) Purchases", pl["purchases"], None),
        ("(ii) Carriage Inward", pl["carriage_inward"], None),
        ("(iii) Sub-Total", [p + c for p, c in zip(pl["purchases"], pl["carriage_inward"])], "total"),
        ("(iv) Add Opening Stock", pl["opening_stock"], None),
        ("(vi) Less Closing Stock", pl["closing_stock"], None),
        ("(vii) Total Cost of Sales", pl["cost_of_sales"], "section"),
        ("3. Operating Expenses", None, "section"),
        ("Salary", pl["salary"], None),
        ("Rent", pl["rent"], None),
        ("Power & Fuel", pl["power_fuel"], None),
        ("Travelling & Conveyance", pl["travelling"], None),
        ("Telephone & Internet", pl["telephone"], None),
        ("Office Expenses", pl["office"], None),
        ("Printing & Stationery", pl["printing"], None),
        ("Repairs & Maintenance", pl["repairs"], None),
        ("Other Operating Expenses", pl["other_operating"], None),
        ("4. Operating Profit (Before Interest & Depreciation)", pl["op_profit_before_int_dep"], "section"),
        ("5. Interest on Cash Credit / OD", pl["interest_cc"], "section"),
        ("6. Depreciation", pl["depreciation"], "section"),
        ("7. Operating Profit (After Interest & Depreciation)", pl["op_profit_after_int_dep"], "section"),
        ("9. Profit Before Tax", pl["pbt"], "section"),
        ("10. Provision for Tax", pl["tax"], "subsection"),
        ("11. Net Profit", pl["net_profit"], "section"),
        ("12. Dividend", pl["dividend"], "subsection"),
        ("13. Retained Profit", pl["retained_profit"], "total"),
    ]
    _write_table(ws_pl, "PROJECTED PL", pl_rows)

    ws_bs = wb.create_sheet("PROJECTED BS")
    bs_rows = [
        ("CURRENT LIABILITIES", None, "section"),
        ("Short Term Borrowings from Applicant Bank", bs["cc_from_applicant_bank"], None),
        ("Sundry Creditors (Trade)", bs["sundry_creditors"], None),
        ("Other Current Liabilities", bs["other_current_liabilities"], None),
        ("Total Current Liabilities (B)", bs["total_current_liabilities"], "total"),
        ("TERM LIABILITIES", None, "section"),
        ("Term Loans", bs["term_loans"], None),
        ("Total Term Liabilities (C)", bs["term_loans"], "total"),
        ("TOTAL OUTSIDE LIABILITIES (D)", bs["total_outside_liabilities"], "section"),
        ("SHAREHOLDERS FUNDS", None, "section"),
        ("Share Capital", bs["share_capital"], None),
        ("Surplus / Deficit in P&L", bs["surplus_pl"], None),
        ("Net Worth (E)", bs["net_worth"], "total"),
        ("TOTAL LIABILITIES (F)", bs["total_liabilities"], "section"),
        ("CURRENT ASSETS", None, "section"),
        ("Cash & Bank Balances", bs["cash_bank"], None),
        ("Receivables", bs["receivables"], None),
        ("Stocks in Trade", bs["stocks"], None),
        ("Other Current Assets", bs["other_current_assets"], None),
        ("Total Current Assets (G)", bs["total_current_assets"], "total"),
        ("FIXED ASSETS", None, "section"),
        ("Gross Block", bs["gross_block"], None),
        ("Depreciation to Date", bs["dep_to_date"], None),
        ("Net Block (H)", bs["net_block"], "total"),
        ("TOTAL ASSETS (J)", bs["total_assets"], "section"),
        ("NET WORKING CAPITAL", bs["net_working_capital"], "total"),
    ]
    _write_table(ws_bs, "PROJECTED BS", bs_rows)
    ws_bs.append(["Balance Status"] + bs["balance_status"])
    for c in range(1, 7):
        cell = ws_bs.cell(row=ws_bs.max_row, column=c)
        cell.border = THIN
        cell.fill = TOTAL_FILL
        if c == 1:
            cell.font = Font(bold=True)

    ws_wc = wb.create_sheet("WORKING CAPITAL ANALYSIS")
    wc_rows = [
        ("A. CURRENT ASSETS", None, "section"),
        ("Raw Material", wc["raw_material"], None),
        ("Work in Progress", wc["wip"], None),
        ("Finished Goods", wc["finished_goods"], None),
        ("Receivables / Sundry Debtors", wc["receivables"], None),
        ("Cash & Bank", wc["cash_bank"], None),
        ("Other Current Assets", wc["other_current_assets"], None),
        ("Total Current Assets (A)", wc["total_current_assets"], "total"),
        ("B. CURRENT LIABILITIES (OTHER THAN BANK)", None, "section"),
        ("Sundry Creditors", wc["sundry_creditors"], None),
        ("Outstanding Expenses", wc["outstanding_expenses"], None),
        ("Statutory Liabilities", wc["statutory_liabilities"], None),
        ("Total Current Liabilities (B)", wc["total_current_liabilities"], "total"),
        ("C. WORKING CAPITAL GAP (A - B)", wc["wc_gap"], "section"),
        ("D. BORROWER CONTRIBUTION (25% of CA)", wc["borrower_contribution"], "subsection"),
        ("E. MAXIMUM PERMISSIBLE BANK FINANCE (MPBF)", wc["mpbf"], "total"),
        ("F. PROPOSED CC LIMIT", wc["proposed_cc_limit"], "section"),
        ("Alternative - Projected Annual Turnover", wc["projected_annual_turnover"], "subsection"),
        ("Working Capital Requirement @25%", wc["nayak_wc_requirement"], None),
        ("Borrower Contribution @5%", wc["nayak_borrower_contribution"], None),
        ("Eligible Bank Finance @20%", wc["nayak_eligible_bank_finance"], "total"),
    ]
    _write_table(ws_wc, "WORKING CAPITAL ANALYSIS", wc_rows)

    ws_dep = wb.create_sheet("DEPRECIATION SCHEDULE")
    dep_rows = [
        ("Opening Gross Block", pl["dep_opening_gross_block"], None),
        ("Additions During Year", pl["dep_additions"], None),
        ("Gross Block", pl["dep_gross_block"], "total"),
        ("Opening Accumulated Depreciation", pl["dep_opening_accumulated"], None),
        ("Depreciation for the Year", pl["depreciation"], "section"),
        ("Accumulated Depreciation", pl["dep_accumulated"], "total"),
        ("Closing Net Block", pl["dep_net_block"], "section"),
    ]
    _write_table(ws_dep, "DEPRECIATION SCHEDULE - PRINT FORMAT", dep_rows)

    ws_ratios = wb.create_sheet("FINANCIAL RATIOS ANALYSIS")
    ws_ratios.append(["S.No", "Particulars", "Numerator", "Denominator", *YEARS, "Bank Acceptable Benchmark", "Status"])
    for cell in ws_ratios[1]:
        cell.fill = DEFAULT_HEADER_FILL
        cell.font = Font(bold=True, color="0F172A")
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center")

    for row in ratios:
        ws_ratios.append([
            row["S.No"], row["Particulars"], row["Numerator"], row["Denominator"],
            row["FY1"], row["FY2"], row["FY3"], row["FY4"], row["FY5"],
            row["Bank Acceptable Benchmark"], row["Status"],
        ])
    for r in ws_ratios.iter_rows(min_row=2):
        for c in r:
            c.border = THIN
    ws_ratios.column_dimensions["B"].width = 30
    ws_ratios.column_dimensions["C"].width = 22
    ws_ratios.column_dimensions["D"].width = 22
    for c in "EFGHI":
        ws_ratios.column_dimensions[c].width = 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
